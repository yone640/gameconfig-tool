# -*- coding: utf-8 -*-
"""Excel loader: read a config workbook into a list of dict rows."""
from __future__ import annotations

import os
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook


@dataclass
class SheetData:
    """One sheet of a workbook, as rows of {column: value}."""
    name: str
    columns: List[str]
    rows: List[Dict[str, Any]]


@dataclass
class WorkbookData:
    """A whole config table file."""
    path: str
    sheets: List[SheetData] = field(default_factory=list)

    def sheet(self, name: str) -> Optional[SheetData]:
        for s in self.sheets:
            if s.name == name:
                return s
        return None


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str):
        return value.strip() == ""
    return False


def load_excel(path: str, header_row: int = 1) -> WorkbookData:
    """Load an .xlsx file.

    header_row is 1-based; the header row gives column names. Rows with all
    blank cells are skipped. Merged cells are read via the top-left value.
    """
    if not os.path.exists(path):
        raise FileNotFoundError(path)
    wb = load_workbook(path, data_only=True, read_only=True)
    wbdata = WorkbookData(path=path)
    for ws in wb.worksheets:
        rows_iter = ws.iter_rows(values_only=True)
        rows = list(rows_iter)
        if not rows:
            wbdata.sheets.append(SheetData(name=ws.title, columns=[], rows=[]))
            continue
        header_idx = header_row - 1
        if header_idx >= len(rows):
            wbdata.sheets.append(SheetData(name=ws.title, columns=[], rows=[]))
            continue
        header = [str(c).strip() if c is not None else "" for c in rows[header_idx]]
        data_rows = rows[header_idx + 1:]
        records: List[Dict[str, Any]] = []
        for r in data_rows:
            if all(_is_blank(c) for c in r):
                continue
            record = {}
            for col, val in zip(header, r):
                if col == "":
                    continue
                record[col] = val
            records.append(record)
        wbdata.sheets.append(SheetData(name=ws.title, columns=header, rows=records))
    wb.close()
    return wbdata


def load_directory(directory: str, header_row: int = 1) -> List[WorkbookData]:
    """Load every .xlsx under directory (non-recursive)."""
    out = []
    for name in sorted(os.listdir(directory)):
        if name.lower().endswith((".xlsx", ".xlsm")):
            out.append(load_excel(os.path.join(directory, name), header_row=header_row))
    return out
